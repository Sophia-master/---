from typing import Dict

import telebot
from openpyxl.workbook import Workbook
from telebot import types
from telebot.types import ReplyKeyboardMarkup, InlineKeyboardMarkup

from credit_form import CreditForm

bot = telebot.TeleBot('6988662541:AAF1Mzu3fNr1okYiHUfOxOcUVTfeNwXIihU')  # подключаем бота к коду

credit_forms: Dict[int, CreditForm] = {}


def get_or_create_credit_form(chat_id: int) -> CreditForm:  # используем созданный класс
    credit_form = credit_forms.get(chat_id)
    if not credit_form:
        credit_form = CreditForm()
        credit_forms[chat_id] = credit_form
    return credit_form


def get_btn_ready() -> ReplyKeyboardMarkup:  # функция создающая кнопки да и нет
    markup = types.ReplyKeyboardMarkup()
    btn_yes = types.KeyboardButton('Да')
    btn_no = types.KeyboardButton('Нет')
    markup.row(btn_yes, btn_no)

    return markup


def get_btn_restart() -> ReplyKeyboardMarkup:  # функция запускающая бота с начала
    markup = types.ReplyKeyboardMarkup()
    btn_restart = types.KeyboardButton('Начать с начала')
    markup.row(btn_restart)

    return markup


@bot.message_handler(commands=['start'])  # обработчик команды /start
def handler_wellcome(message):  # функция приветствующая гостя
    markup = get_btn_ready()  # создание кнопок

    bot.send_message(message.chat.id, f'Привет, {message.from_user.first_name}. Вы готовы начать?', reply_markup=markup)
    bot.register_next_step_handler(message, handler_start)


def handler_start(message):  # функция обрабатывающая ответ на вопрос о готовности
    if message.text == 'Да':
        bot.send_message(message.chat.id, 'Введите срок кредитования (в месяцах)', reply_markup=types.ReplyKeyboardRemove())
        bot.register_next_step_handler(message, handle_duration_month)
    elif message.text == 'Нет':
        markup = get_btn_restart()
        bot.send_message(message.chat.id, 'Спасибо за то, что воспользовались нашими услугами, рады помочь!',
                         reply_markup=markup)

        bot.register_next_step_handler(message, handler_wellcome)


def handle_duration_month(message):  # функция обрабатывающая введенный month
    credit_form = get_or_create_credit_form(message.chat.id)
    try:
        credit_form.duration_month = int(message.text)
        bot.send_message(message.chat.id, 'Введите процентную ставку (процент в год) Например: 2.4')
        bot.register_next_step_handler(message, handle_percent)
    except ValueError:
        bot.send_message(message.chat.id, 'Это не число! Попробуйте ещё раз')
        bot.register_next_step_handler(message, handle_duration_month)


def handle_percent(message):  # функция обрабатывающая введенный percent
    credit_form = get_or_create_credit_form(message.chat.id)
    try:
        credit_form.percent = float(message.text)
        bot.send_message(message.chat.id, 'Введите предполагаемую сумму кредита')
        bot.register_next_step_handler(message, handle_amount)
    except ValueError:
        bot.send_message(message.chat.id, 'Это не число! Попробуйте ещё раз')
        bot.register_next_step_handler(message, handle_percent)


def get_btn_payment_type() -> ReplyKeyboardMarkup:  # функция создающая кнопки выбора платежа
    markup = types.ReplyKeyboardMarkup()
    btn_dif = types.KeyboardButton('Дифференцированный платеж')
    markup.row(btn_dif)
    btn_any = types.KeyboardButton('Аннуитетный платеж')
    markup.row(btn_any)
    btn_question = types.KeyboardButton('А в чем разница?')
    markup.row(btn_question)

    return markup

# def get_btn_payment_q() -> InlineKeyboardMarkup:
#     markup = types.InlineKeyboardMarkup()
#     btn_dif = types.InlineKeyboardButton('Что такое дифференцированный платеж?')
#     markup.row(btn_dif)
#     btn_any = types.InlineKeyboardButton('Что такое аннуитетный платеж?')
#     markup.row(btn_any)
#     return markup

def handle_amount(message):  # функция обрабатывающая введенный amount
    credit_form = get_or_create_credit_form(message.chat.id)
    try:
        credit_form.amount = float(message.text)

        markup = get_btn_payment_type()
        bot.send_message(message.chat.id, 'Выберите вид платежа', reply_markup=markup)

        bot.register_next_step_handler(message, handle_payment_type)
    except ValueError:
        bot.send_message(message.chat.id, 'Это не число! Попробуйте ещё раз')
        bot.register_next_step_handler(message, handle_amount)



def call_differentiated(month, percent, amount):  # функция отвечающая за расчет при дифференцированном платеже
    percent /= 100
    percent /= 12
    remains = amount / month  # ежемесячный платеж по основному долгу
    residual = amount  # остаточная задолженность
    payment_total = 0
    payment_schedule = []
    for i in range(month):
        residual -= remains
        monthly_payment = remains + (residual * percent)  # ежемесячный платеж
        payment_schedule.append(monthly_payment)
        payment_total += monthly_payment  # платеж за весь период
    payment_over = payment_total - amount  # переплата
    return payment_schedule, payment_total, payment_over


def call_annuitant(month, percent, amount):  # функция отвечающая за расчет при аннуитетном платеже
    percent /= 100
    percent /= 12
    k = (percent * ((1 + percent) ** month)) / (((1 + percent) ** month) - 1)  # коэффициент аннуитета
    payment_total = 0
    payment_schedule = []
    for i in range(month):  #
        monthly_payment = k * amount  # ежемесячный платеж
        payment_schedule.append(monthly_payment)
        payment_total += monthly_payment  # платеж за весь период
    payment_over = payment_total - amount  # переплата
    return payment_schedule, payment_total, payment_over


def create_file(chat_id: int, payment_schedule: list) -> str:  # функция структурирующая полученные данные
    wb = Workbook()
    ws = wb.active
    ws[f'A1'].value = f'месяц'
    ws[f'B1'].value = f'платёж (руб)'
    for i, payment in enumerate(payment_schedule):
        ws[f'A{i + 2}'].value = i + 1
        ws[f'B{i + 2}'].value = payment
    file_name = f'calc_{chat_id}.xlsx'
    wb.save(file_name)
    return file_name


def send_result(chat_id: int, file_name: str, payment_total, payment_over):  # функция составляющая сообщение
    msg = f'Платеж за весь период составит {payment_total} руб\n'
    msg += f'Переплата будет {payment_over} руб'

    bot.send_message(chat_id, msg)
    bot.send_document(
        chat_id=chat_id,
        caption='ежемесячный платёж',
        document=open(file_name, 'rb'),
    )


def differentiated(message, month, percent, amount):  # функция для дифференцированного платежа
    payment_schedule, payment_total, payment_over = call_differentiated(month, percent, amount)
    file_name = create_file(message.chat.id, payment_schedule)
    send_result(message.chat.id, file_name, payment_total, payment_over)


def annuitant(message, month, percent, amount):  # функция для аннуитетного платежа
    payment_schedule, payment_total, payment_over = call_annuitant(month, percent, amount)
    file_name = create_file(message.chat.id, payment_schedule)
    send_result(message.chat.id, file_name, payment_total, payment_over)


def handle_payment_type(message):  # функция отправляющая сообщение
    credit_form = get_or_create_credit_form(message.chat.id)
    if message.text == 'Дифференцированный платеж':
        differentiated(message, credit_form.duration_month, credit_form.percent, credit_form.amount)

        markup = get_btn_restart()
        bot.send_message(message.chat.id, 'Спасибо за то, что воспользовались нашими услугами, рады помочь!',
                         reply_markup=markup)

        bot.register_next_step_handler(message, handler_wellcome)
    elif message.text == 'Аннуитетный платеж':
        annuitant(message, credit_form.duration_month, credit_form.percent, credit_form.amount)

        markup = get_btn_restart()
        bot.send_message(message.chat.id, 'Спасибо за то, что воспользовались нашими услугами, рады помочь!',
                         reply_markup=markup)

        bot.register_next_step_handler(message, handler_wellcome)
    elif message.text == 'А в чем разница?':
        markup = get_btn_payment_type()
        bot.send_message(message.chat.id, 'При аннуитетном платеже проценты начисляются на всю сумму кредита, '
                                          'а после делится на срок кредитования. В итоге заемщик каждый месяц '
                                          'выплачивает одну и ту же сумму, до тех пор, пока не погасит задолженность')
        bot.send_message(message.chat.id, 'Дифференцированный платеж предполагает начисление процентов на остаток '
                                          'задолженности, поэтому состоит из двух частей. Первой — платежу по основной '
                                          'задолженности и второй — платежу по процентам. Так как с каждым месяцем '
                                          'основная задолженность будет уменьшаться, то и платеж по процентам будет '
                                          'уменьшаться. Поэтому при дифференцированном платеже сумма ежемесячных выплат'
                                          'разная и с каждым месяцем она все меньше.', reply_markup=markup)

        bot.register_next_step_handler(message, handle_payment_type)


bot.infinity_polling()  # запускает цикл обработки событий бота
